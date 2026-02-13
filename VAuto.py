import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configurações de Design do Sistema
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class AuditProcessor:
    def __init__(self, file_path, et_value=100000):
        self.file_path = file_path
        self.et_value = et_value
        self.keywords = ['ajuste', 'estorno', 'erro', 'manual', 'urgente', 'socio', 'conforme']

    def process_audit(self, output_path):
        # Carregamento de dados (CSV ou Excel)
        if self.file_path.endswith('.csv'):
            df = pd.read_csv(self.file_path)
        else:
            df = pd.read_excel(self.file_path, engine='openpyxl')
        
        # Padronização de colunas (trata colunas sem nome como 'Historico')
        df.columns = [c if 'Unnamed' not in str(c) else 'Historico' for c in df.columns]
        colunas_originais = list(df.columns)

        # Tratamento de tipos e criação da coluna de valor bruto
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
        df['Débito'] = pd.to_numeric(df['Débito'], errors='coerce').fillna(0)
        df['Crédito'] = pd.to_numeric(df['Crédito'], errors='coerce').fillna(0)
        df['Valor_Bruto'] = df['Débito'] + df['Crédito']

        # --- Procedimentos de Auditoria ---
        df['10x_Media'] = df['Valor_Bruto'] > (df.groupby('Cta.C.Part.')['Valor_Bruto'].transform('mean') * 10)
        df['Excede_ET'] = df['Valor_Bruto'] > self.et_value
        df['Redondo'] = df['Valor_Bruto'].apply(lambda x: x > 0 and x % 100 == 0)
        df['Sem_Hist'] = df['Historico'].apply(lambda x: len(str(x)) < 2 or pd.isna(x))
        df['Fds'] = df['Data'].dt.dayofweek.isin([5, 6])
        df['Palavra_Chave'] = df['Historico'].str.contains('|'.join(self.keywords), case=False, na=False)
        
        procedimentos = {
            "Geral": None,
            "10xMedia": "10x_Media",
            "ExcedeET": "Excede_ET",
            "Redondo": "Redondo",
            "Sem Historico": "Sem_Hist",
            "Final De Semana": "Fds",
            "Palavras Chave": "Palavra_Chave"
        }

        # Geração do Excel com pd.ExcelWriter
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for nome_aba, coluna_filtro in procedimentos.items():
                if nome_aba == "Geral":
                    df.to_excel(writer, sheet_name="Geral", index=False, startrow=7)
                else:
                    colunas_finais = colunas_originais + [coluna_filtro]
                    colunas_existentes = [c for c in colunas_finais if c in df.columns]
                    df[colunas_existentes].to_excel(writer, sheet_name=nome_aba, index=False, startrow=7)
                
                # Chama a função de estilo para cada aba criada
                self.aplicar_estilo(writer, nome_aba)

        # Estatísticas para o Dashboard
        stats = {
            '10x Média': df['10x_Media'].sum(),
            'Excede ET': df['Excede_ET'].sum(),
            'Vlr Redondo': df['Redondo'].sum(),
            'Sem Histórico': df['Sem_Hist'].sum(),
            'Fim de Semana': df['Fds'].sum(),
            'Palavras-Chave': df['Palavra_Chave'].sum()
        }
        return output_path, stats
    
    def aplicar_estilo(self, writer, nome_aba):
        ws = writer.sheets[nome_aba]
        
        max_col = ws.max_column
        ultima_col_letra = get_column_letter(max_col)

        # Definição de Fontes e Cores 
        fonte_base = Font(name='Arial', size=10)
        fonte_negrito = Font(name='Arial', size=10, bold=True)
        fill_cabecalho = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

        # 1. Cabeçalho Superior 1 a 3 (Padronizado)
        ws.merge_cells(f'A1:{ultima_col_letra}1')
        ws['A1'] = "Villela e Associados Auditoria e Consultoria Ltda."
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color="000000")
        
        ws.merge_cells(f'A2:{ultima_col_letra}2')
        ws['A2'] = f"RELATÓRIO DE AUDITORIA - {nome_aba.upper()}"
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        

        # Data por extenso em Português
        meses = {1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 5: "maio", 6: "junho",
                7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"}
        agora = datetime.now()
        data_extenso = f"{agora.day} de {meses[agora.month]} de {agora.year}"
        
        ws.merge_cells(f'A3:{ultima_col_letra}3')
        ws['A3'] = f"Processado em: {data_extenso}"
        ws['A3'].font = Font(name='Arial', size=10, italic=True)
        
        ## --- Seção de procedimentos ( linha 4 a 7) ---

            # Seção de Objetivo
        objetivos_proc = {
            "Geral": "Apresenta a listagem completa de todos os lançamentos processados no período.",
            "10xMedia": "Identificar outliers (valores muito acima do padrão da conta específica).",
            "ExcedeET": "Filtrar lançamentos que possuem relevância financeira, acima da materialidade (Erro Tolerável) definida.",
            "Redondo": "Identificar lançamentos com valores redondos (possíveis estimativas).",
            "Sem Historico": "Detectar lançamentos com descrições ausentes ou curtas.",
            "Final De Semana": "Filtrar lançamentos realizados em sábados ou domingos, o que pode sugerir lançamentos retroativos ou falta de controle de acesso ao sistema.",
            "Palavras Chave": "Buscar termos sensíveis (ajuste, estorno, erro, etc) no histórico."
        }
                # -- linha 4 e 5 --
        ws['A4'] = "Objetivo:"
        ws['A4'].font = fonte_negrito
        ws.merge_cells(f'A5:{ultima_col_letra}5')
        ws['A5'] = objetivos_proc.get(nome_aba, "Análise de integridade contábil.")
        ws['A5'].font = fonte_base
        ws['A5'].alignment = Alignment(wrap_text=True, vertical="top")


        proc_feitos = {
            "Geral": "Apenas listagem completa com todos procedimentos aplicados.",
            "10xMedia": "Ele verifica se o valor de cada lançamento individual é 10 vezes maior que a média histórica daquela conta específica.",
            "ExcedeET": "O código compara a coluna Valor_Bruto com a variável self.et_value (o Erro Tolerável que você digita na interface). Se Valor_Bruto > ET, ele marca como positivo.",
            "Redondo": "Ele verifica se o valor é maior que zero e se o resto da divisão por 100 é igual a zero",
            "Sem Historico": "Código marca lançamentos onde o campo 'Histórico' tem menos de 2 caracteres (len < 2) ou está vazio (pd.isna).",
            "Final De Semana": "Ele converte a coluna 'Data' para o formato datetime e verifica o índice do dia. No Python, 5 é Sábado e 6 é Domingo.",
            "Palavras Chave": "O código varre a coluna ''Histórico'' procurando por: 'ajuste', 'estorno', 'erro', 'manual', 'urgente', 'socio', 'conforme'.",
        }

                # -- linha 6 e 7 --
        ws['A6'] = "Procedimento Feito:"
        ws['A6'].font = fonte_negrito
        ws.merge_cells(f'A7:{ultima_col_letra}7')
        ws['A7'] = proc_feitos.get(nome_aba, "Aplicação de procedimentos para análise de riscos.")
        ws['A7'].font = fonte_base
        ws.row_dimensions[7].height = 26.85
        ws['A7'].alignment = Alignment(wrap_text=True, vertical="top")

        # Usamos range numérico para evitar conflitos com células mescladas acima
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            
            # Largura fixa de 13 (aprox. 95 pixels)
            ws.column_dimensions[col_letter].width = 13



            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 16 
                
                # Cabeçalho da Tabela (Linha 8) - Fundo #A6A6A6
                cell_header = ws.cell(row=8, column=col_idx)
                cell_header.fill = fill_cabecalho
                cell_header.font = fonte_negrito
                cell_header.alignment = Alignment(horizontal="center", vertical="center")

                # Nome da coluna para identificar onde aplicar formatos específicos
                col_name = cell_header.value

                # Formatar dados das células (Linha 9 em diante)
                for row_idx in range(9, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = fonte_base
                    
                    # A. FORMATO DE DATA (Coluna 1 / Coluna A)
                    if col_idx == 1:
                        cell.number_format = 'DD/MM/YYYY'
                    
                    # B. FORMATO DE MOEDA (Colunas de valores financeiros)
                    valores_financeiros = ['Débito', 'Crédito', 'Saldo-Exercicio', 'Valor_Bruto']
                    if col_name in valores_financeiros:
                        cell.number_format = '"R$ " #,##0.00'
                        cell.alignment = Alignment(horizontal="left")
                    # C. FORMATO DE TEXT CENTER
                    valores_num = ['Cta.C.Part.', 'Número', 'Data' ]
                    if col_name in valores_num:
                        cell.alignment = Alignment(horizontal="center")


class DashboardWindow(ctk.CTkToplevel):
    def __init__(self, stats):
        super().__init__()
        self.title("Painel de Análise de Riscos")
        self.geometry("800x600")
        
        label = ctk.CTkLabel(self, text="Resumo de Ocorrências por Procedimento", font=("Roboto", 22, "bold"))
        label.pack(pady=20)

        # Gráfico de Barras Centralizado
        fig, ax = plt.subplots(figsize=(8, 5))
        fig.patch.set_facecolor('#1e1e1e')
        ax.set_facecolor('#2b2b2b')
        
        names = list(stats.keys())
        values = list(stats.values())
        
        bars = ax.bar(names, values, color='#1f538d')
        ax.set_xticklabels(names, rotation=45, ha='right', color='white')
        ax.tick_params(axis='y', colors='white')
        
        # Adiciona rótulos numéricos nas barras
        for bar in bars:
            yval = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2, yval + 0.1, int(yval), ha='center', color='white', fontweight='bold')

        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=20, pady=20)

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Auditor Contábil Pro")
        self.geometry("500x550")
        self.configure(fg_color="#121212") 

        self.main_container = ctk.CTkFrame(self, fg_color="#1e1e1e", corner_radius=25)
        self.main_container.pack(pady=40, padx=40, fill="both", expand=True)

        self.label = ctk.CTkLabel(self.main_container, text="Auditoria Contábil", font=("Roboto", 27, "bold"))
        self.label.pack(pady=(35, 20))

        self.et_label = ctk.CTkLabel(self.main_container, text="Valor do Erro Tolerável (ET):")
        self.et_label.pack(pady=(10, 5))
        
        self.et_entry = ctk.CTkEntry(self.main_container, width=220, justify="center")
        self.et_entry.insert(0, "100000")
        self.et_entry.pack(pady=10)

        self.btn_run = ctk.CTkButton(self.main_container, text="Selecionar Razão (.xlsx)", command=self.run_process)
        self.btn_run.pack(pady=25)

        #PARTE BOTAO INVISIVEL
            # --- botao do dashboard -- 
        self.btn_dash = ctk.CTkButton(self.main_container, text="Ver Painel de Análise", command=self.open_dashboard,
                                    fg_color="transparent", border_width=2)
            # --- botao para abrir resultado ---
        self.btn_result = ctk.CTkButton(self.main_container, text="Abrir Resultado", command=self.open_result,
                                    fg_color="transparent", border_width=2)

    def open_result(self):
        path_saida = self.ultimo_resultado
        if os.path.exists(path_saida):
            os.startfile(path_saida)
        else:
            messagebox.showerror("Erro", "Arquivo de saída não encontrado.")

    def run_process(self):
        path_entrada = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
        if path_entrada:
            path_saida = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Razao_Auditado_Final.xlsx")
            if path_saida:
                try:
                    et = float(self.et_entry.get())
                    proc = AuditProcessor(path_entrada, et)
                    _, self.stats = proc.process_audit(path_saida)

                    self.ultimo_resultado = path_saida

                    messagebox.showinfo("Sucesso", "Arquivo processado com sucesso!")

                    self.btn_dash.pack(pady=10)
                    self.btn_result.pack(pady=12)
                except Exception as e:
                    messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

    def open_dashboard(self):
        if hasattr(self, 'stats'):
            DashboardWindow(self.stats)

if __name__ == "__main__":
    app = App()
    app.mainloop()