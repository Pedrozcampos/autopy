from csv import writer
from datetime import datetime   
from tkinter.font import Font
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
from openpyxl.styles import Font, PatternFill, Alignment

# Configurações de Design
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class AuditProcessor:
    def __init__(self, file_path, et_value=100000):
        self.file_path = file_path
        self.et_value = et_value
        self.keywords = ['ajuste', 'estorno', 'erro', 'manual', 'urgente', 'socio', 'conforme']

    def process_audit(self, output_path):
        # Carregamento (CSV ou Excel)
        if self.file_path.endswith('.csv'):
            df = pd.read_csv(self.file_path)
        else:
            df = pd.read_excel(self.file_path, engine='openpyxl')
        

        # Nota: O pandas as vezes lê a coluna sem nome como 'Unnamed: 2'
        df.columns = [c if 'Unnamed' not in str(c) else 'Historico' for c in df.columns]
        Colunas_originais = list(df.columns)

        df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
        df['Débito'] = pd.to_numeric(df['Débito'], errors='coerce').fillna(0)
        df['Crédito'] = pd.to_numeric(df['Crédito'], errors='coerce').fillna(0)
        df['Valor_Bruto'] = df['Débito'] + df['Crédito']

        # --- Procedimentos ---
        
        df['10x_Media'] = df['Valor_Bruto'] > (df.groupby('Cta.C.Part.')['Valor_Bruto'].transform('mean') * 10)
        df['Excede_ET'] = df['Valor_Bruto'] > self.et_value
        df['Redondo'] = df['Valor_Bruto'].apply(lambda x: x > 0 and x % 100 == 0)
        df['Sem_Hist'] = df['Historico'].apply(lambda x: len(str(x)) < 5 or pd.isna(x))
        df['Fds'] = df['Data'].dt.dayofweek.isin([5, 6])
        df['Palavra_Chave'] = df['Historico'].str.contains('|'.join(self.keywords), case=False, na=False)
        
        Procedimentos = {
            "10xMedia": "10x_Media",
            "ExcedeET": "Excede_ET",
            "Redondo": "Redondo",
            "Sem Historico": "Sem_Hist",
            "Final De Semana": "Fds",
            "Palavras Chave": "Palavra_Chave"
        }
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Geral", index=False, startrow=7)
            self.aplicar_estilo_profissional(writer, "Geral")
            # Abas Específicas:
            for nome_aba, coluna_filtro in Procedimentos.items():
                df_aba = df.copy()

                colunas_finais = Colunas_originais + [coluna_filtro]
                
                # Garantimos que só pegamos colunas que realmente existem (evita erros)
                colunas_existentes = [c for c in colunas_finais if c in df_aba.columns]
                
                df_aba[colunas_existentes].to_excel(writer, sheet_name=nome_aba, index=False, startrow=3)
                self.aplicar_estilo_profissional(writer, nome_aba)

        output_path = "Razao_Auditado_Final.xlsx"
        df.to_excel(output_path, index=False)
        
        # Gerar estatísticas para o Painel
        stats = {
            '10x Média': df['10x_Media'].sum(),
            'Excede ET': df['Excede_ET'].sum(),
            'Vlr Redondo': df['Redondo'].sum(),
            'Sem Histórico': df['Sem_Hist'].sum(),
            'Fim de Semana': df['Fds'].sum(),
            'Palavras-Chave': df['Palavra_Chave'].sum()
        }
        return output_path, stats
    
    def aplicar_estilo(Self, writer, nome_aba):
        # --- TEXTO PADRÃO NAS PRIMEIRAS LINHAS ---
            # Mescla as células da coluna A até a última coluna de dados
            ws = writer.sheets[nome_aba]
            ultima_col_letra = ws.cell(row=4, column=ws.max_column).column_letter

            ws.merge_cells(f'A1:{ultima_col_letra}1')
            ws['A1'] = "Villela e Associados Auditoria e Consultoria Ltda."
            ws['A1'].font = Font(size=16, bold=True, color="1F4E78")

            ws.merge_cells(f'A2:{ultima_col_letra}2')
            ws['A2'] = "CLIENTE MODELO"
            ws['A2'].font = Font(size=13, italic=True)

            meses = {
            1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 
            5: "maio", 6: "junho", 7: "julho", 8: "agosto", 
            9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"
            }   
            agora = datetime.now()
            data_extenso = f"{agora.day} de {meses[agora.month]} de {agora.year}"
            ws.merge_cells(f'A3:{ultima_col_letra}3')
            ws['A3'] = f"{data_extenso}"
            ws['A3'].font = Font(size=11, italic=True)
            objetivos_proc = {
            "Geral": "Apresenta a listagem completa de todos os lançamentos processados no período.",
            "10xMedia": "Identifica outliers (valores muito acima do padrão da conta específica).",
            "ExcedeET": "Filtra lançamentos acima da Materialidade (Erro Tolerável) definida pelo usuário.",
            "Redondo": "Identifica lançamentos com valores redondos, que podem indicar estimativas ou falta de precisão.",
            "Sem Historico": "Detecta lançamentos com descrições ausentes ou excessivamente curtas.",
            "Final De Semana": "Filtra lançamentos realizados em sábados ou domingos (dias não úteis).",
            "Palavras Chave": "Busca termos sensíveis no histórico que podem indicar ajustes, erros ou fraudes."
            }

            ws.merge_cells(f'A4:{ultima_col_letra}4')
            ws['A4'] = "Objetivo:"
            texto_objetivo = objetivos_proc.get(nome_aba, "Análise de integridade contábil.")
            ws.merge_cells(f'A5:{ultima_col_letra}5')
            ws['A5'] = texto_objetivo
            ws['A5'].font = Font(size=11, italic=True)
        # -- cabeçalho e dados ---
            fonte_padrao = Font(name='Arial', size=10)
            fonte_cabecalho = Font(name='Arial', size=10, bold=True)

            fill_cabecalho = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
            for cell in ws[7]:
                cell.fill = fill_cabecalho
                cell.font = fonte_cabecalho
                cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in ws.columns:
                    column_letter = col[0].column_letter
                    ws.column_dimensions[column_letter].width = 13
                    
                    # Aplicar a fonte Arial nos dados (da linha 8 em diante)
                    for cell in col[7:]: # Começa da linha 8 (índice 7 do Python)
                        cell.font = fonte_padrao

class DashboardWindow(ctk.CTkToplevel):
    def __init__(self, stats):
        super().__init__()
        self.title("Painel de Análise de Riscos")
        self.geometry("700x500")
        
        label = ctk.CTkLabel(self, text="Resumo de Ocorrências por Procedimento", font=("Roboto", 18, "bold"))
        label.pack(pady=10)

        # Criando o gráfico com Matplotlib
        fig, ax = plt.subplots(figsize=(6, 4), dpi=100)
        fig.patch.set_facecolor('#2b2b2b') # Cor de fundo combinando com Dark Mode
        ax.set_facecolor('#2b2b2b')
        
        names = list(stats.keys())
        values = list(stats.values())
        
        bars = ax.bar(names, values, color='#1f538d')
        ax.set_xticklabels(names, rotation=45, ha='right', color='white', fontsize=10)
        plt.subplots_adjust(bottom=0.25)
        ax.tick_params(axis='y', colors='white')
        ax.spines['bottom'].set_color('white')
        ax.spines['left'].set_color('white')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        plt.tight_layout()

        # Adiciona os números em cima das barras
        for bar in bars:
            yval = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2, yval + 0.1, int(yval), ha='center', color='white')

        canvas = FigureCanvasTkAgg(fig, master=self)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=20, pady=20)


        

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Configurações de Janela
        self.title("Auditor Contábil")
        self.geometry("500x550")
        
        # 1. Cor de Fundo Profunda
        self.configure(fg_color="#121212") 

        # 2. Container Principal
        self.main_container = ctk.CTkFrame(
            self, 
            fg_color="#1e1e1e",      
            corner_radius=25,        
            border_width=1,           
            border_color="#333333"    
        )
        self.main_container.pack(pady=40, padx=40, fill="both", expand=True)

        # Título
        self.label = ctk.CTkLabel(
            self.main_container, 
            text="Auditoria Contábil", 
            font=("Roboto", 27, "bold"),
            text_color="#ffffff"
        )
        self.label.pack(pady=(35, 20))

        # Campo de Entrada ET
        self.et_label = ctk.CTkLabel(self.main_container, text="Valor do Erro Tolerável (ET):", font=("Roboto", 15))
        self.et_label.pack(pady=(10, 5))
        
        self.et_entry = ctk.CTkEntry(
            self.main_container, 
            width=220, 
            height=40,
            border_color="#444444",
            fg_color="#2a2a2a",
            placeholder_text="Ex: 100",
            justify="center"
        )
        self.et_entry.insert(0, "100")
        self.et_entry.pack(pady=10)

        # 3. Botão Principal 
        self.btn_run = ctk.CTkButton(
            self.main_container, 
            text="Selecionar Razão (.xlsx)", 
            command=self.run_process, 
            height=50,
            width=220,
            corner_radius=12,
            fg_color="#3b82f6",       
            hover_color="#2563eb",    
            font=("Roboto", 14, "bold"),
            border_width=2,
            border_color="#60a5fa"    
        )
        self.btn_run.pack(pady=25)

        # 4. Botão de Painel (Invisível)
        self.btn_dash = ctk.CTkButton(
            self.main_container, 
            text="Ver Painel de Análise", 
            command=self.open_dashboard, 
            fg_color="transparent", 
            border_width=2, 
            border_color="#3b82f6",
            text_color="#3b82f6",
            hover_color="#1e293b",
            height=40
        )
        
    def run_process(self):
        path_entrada = filedialog.askopenfilename(
            title="Selecione o arquivo para auditoria",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )

        if path_entrada:
            path_saida = filedialog.asksaveasfilename(
                title="Salvar arquivo auditado como...",
                defaultextension=".xlsx",
                initialfile="Razao_Auditado_Final.xlsx",
                filetypes=[("Excel files", "*.xlsx")]   
            )

            if path_saida:
                try:
                    et = float(self.et_entry.get())
                    proc = AuditProcessor(path_entrada, et)
                    _, self.stats = proc.process_audit(path_saida)
                    
                    
                    mensagem = f"Sucesso! Arquivo processado e salvo em:\n{path_saida}"
                    messagebox.showinfo("Sucesso", mensagem)
                    self.btn_dash.pack(pady=10) # Revela o botão do painel
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro: {str(e)}")

    def open_dashboard(self):
        if self.stats:
            DashboardWindow(self.stats)

if __name__ == "__main__":
    app = App()
    app.mainloop()